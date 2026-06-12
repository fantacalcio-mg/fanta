import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ListoneAdapter:
    """Adapter per il Listone Ufficiale di Fantacalcio.it.

    Scarica il file .xlsx dalle quotazioni ufficiali.
    """

    def __init__(
        self,
        url: str,
        raw_dir: Path = Path("data/raw/listone_ufficiale"),
        user_agent: str = "fanta-engine/0.1 (personal use)",
        timeout_seconds: float = 30.0,
    ):
        self.url = url
        self.raw_dir = Path(raw_dir)
        self.user_agent = user_agent
        self.timeout = timeout_seconds
        self.raw_dir.mkdir(parents=True, exist_ok=True)

    def find_latest_raw(self) -> Path | None:
        """Trova il file xlsx più recente in raw_dir per data.

        Returns:
            Path al file quotazioni.xlsx più recente, o None se non esiste.
        """
        if not self.raw_dir.exists():
            return None

        dated_dirs = sorted(
            (d for d in self.raw_dir.iterdir() if d.is_dir() and d.name.isdigit()),
            reverse=True,
        )

        for day_dir in dated_dirs:
            candidate = day_dir / "quotazioni.xlsx"
            if candidate.exists():
                logger.info(f"Trovato listone: {candidate}")
                return candidate

        return None

    def download(self, force: bool = False) -> Path:
        """Scarica il listone e lo salva in data/raw/listone_ufficiale/YYYYMMDD/.

        Args:
            force: se True, riscarica anche se il file esiste.

        Returns:
            Path del file xlsx scaricato.

        Raises:
            httpx.HTTPError: se il download fallisce.
        """
        today = datetime.now().strftime("%Y%m%d")
        day_dir = self.raw_dir / today
        day_dir.mkdir(parents=True, exist_ok=True)

        output_path = day_dir / "quotazioni.xlsx"

        if output_path.exists() and not force:
            logger.info(f"File già presente: {output_path}, skipping download")
            return output_path

        logger.info(f"Downloading listone da {self.url}...")
        headers = {"User-Agent": self.user_agent}

        response = httpx.get(self.url, headers=headers, timeout=self.timeout)
        response.raise_for_status()

        output_path.write_bytes(response.content)
        logger.info(f"Listone salvato: {output_path}")

        return output_path

    def parse(self, raw_path: Path) -> list[dict[str, Any]]:
        """Parsa il file xlsx e ritorna una lista di record.

        Mappa colonne reali da xlsx al formato staging.

        Args:
            raw_path: path al file xlsx

        Returns:
            Lista di record con campi: nome, squadra, ruolo_classic, qt_iniziale,
            qt_attuale, fvm, data_nascita (nullable)

        Raises:
            ValueError: se il parsing fallisce.
        """
        try:
            wb = load_workbook(raw_path, data_only=True)
            ws = wb.active
            if ws is None:
                raise ValueError("Nessun foglio trovato nel workbook")

            records = []

            # Salta riga 1 (titolo) e riga 2 (header), inizia da riga 3
            for row_idx in range(3, ws.max_row + 1):
                row = tuple(ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1))

                if not any(row):
                    continue

                record = self._map_row(row, row_idx)
                if record:
                    records.append(record)

            logger.info(f"Parsed {len(records)} giocatori da {raw_path}")
            return records

        except Exception as e:
            raise ValueError(f"Failed to parse {raw_path}: {e}") from e

    @staticmethod
    def _map_row(row: tuple, row_idx: int) -> dict[str, Any] | None:
        """Mappa una riga del listone xlsx al record canonico.

        Mapping colonne (confermato da ispezione file reale Fantacalcio 2025-26):
        0: Id (estratto per ID-based join nel storico)
        1: R (Ruolo: P/D/C/A)
        2: RM (skip — ruolo modulo)
        3: Nome
        4: Squadra
        5: Qt.A (Quotazione attuale)
        6: Qt.I (Quotazione iniziale)
        7-10: skip
        11: FVM

        Vedi DECISIONS.md per il mapping completo.
        """
        try:
            if len(row) < 12:
                return None

            # Col 0 (idx 0): Id (estratto per ID-based join)
            listone_id = None
            if row[0]:
                try:
                    listone_id = int(row[0])
                except (ValueError, TypeError):
                    pass

            # Col 1 (idx 1): Ruolo
            ruolo = str(row[1]).strip().upper() if row[1] else None
            if ruolo not in ("P", "D", "C", "A"):
                return None

            # Col 3 (idx 3): Nome
            nome = str(row[3]).strip() if row[3] else None

            # Col 4 (idx 4): Squadra
            squadra = str(row[4]).strip() if row[4] else None

            # Col 5 (idx 5): Qt.A (attuale)
            qt_attuale = None
            if row[5]:
                try:
                    qt_attuale = int(row[5])
                except (ValueError, TypeError):
                    pass

            # Col 6 (idx 6): Qt.I (iniziale)
            qt_iniziale = None
            if row[6]:
                try:
                    qt_iniziale = int(row[6])
                except (ValueError, TypeError):
                    pass

            # Col 12 (idx 11): FVM
            fvm = None
            if len(row) > 11 and row[11]:
                try:
                    fvm = int(row[11])
                except (ValueError, TypeError):
                    pass

            if not all([nome, squadra, qt_iniziale is not None, qt_attuale is not None]):
                return None

            return {
                "id": listone_id,
                "nome": nome,
                "squadra": squadra,
                "ruolo_classic": ruolo,
                "qt_iniziale": qt_iniziale,
                "qt_attuale": qt_attuale,
                "fvm": fvm,
                "data_nascita": None,  # Non disponibile nel file
            }

        except (ValueError, IndexError, TypeError) as e:
            logger.warning(f"Failed to parse row {row_idx}: {e}")
            return None
