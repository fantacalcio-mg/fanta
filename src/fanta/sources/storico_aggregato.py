"""Adapter per il Storico Aggregato da Fantacalcio.it (XLSX)."""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class StoricoAdapter:
    """Scraper per statistiche aggregate Fantacalcio.it."""

    BASE_URL = "https://www.fantacalcio.it/statistiche-serie-a/{stagione}/italia"

    def __init__(
        self,
        raw_dir: Path = Path("data/raw/storico_aggregato"),
        user_agent: str = "fanta-engine/0.1 (personal use)",
        timeout_seconds: float = 30.0,
    ):
        self.raw_dir = Path(raw_dir)
        self.user_agent = user_agent
        self.timeout = timeout_seconds
        self.raw_dir.mkdir(parents=True, exist_ok=True)

    def download(self, stagione: str, force: bool = False) -> Path:  # noqa: ARG002
        """Scarica XLSX delle statistiche per una stagione.

        Nota: download automatico non implementato. L'utente scarica manualmente
        da Fantacalcio.it e salva in data/raw/storico_aggregato/{stagione}/stats.xlsx

        Args:
            stagione: string format '2024-25'
            force: riscarica anche se esiste (N/A per download manuale)

        Returns:
            Path al file XLSX (se esiste)

        Raises:
            FileNotFoundError: se il file non esiste
        """
        day_dir = self.raw_dir / stagione
        output_path = day_dir / "stats.xlsx"

        if not output_path.exists():
            raise FileNotFoundError(
                f"File {output_path} non trovato. "
                f"Scaricalo manualmente da Fantacalcio.it e salvalo in questa cartella."
            )

        logger.info(f"Storico {stagione} trovato: {output_path}")
        return output_path

    def parse(self, raw_path: Path, stagione: str) -> list[dict[str, Any]]:
        """Parsa XLSX e estrae statistiche giocatori.

        Struttura XLSX:
        - Riga 1: Titolo (skip)
        - Riga 2: Header (18 colonne)
        - Riga 3+: Dati giocatori

        Args:
            raw_path: path al file XLSX
            stagione: stagione (es. '2024-25')

        Returns:
            Lista di dict con campi: nome, squadra, stagione, presenze_voto, media_voto,
            media_fv_fonte, gol, assist_tot, ammonizioni, espulsioni, rig_segnati, rig_parati,
            rig_sbagliati, gol_subiti, autogol

        Raises:
            ValueError: se il parsing fallisce
        """
        try:
            wb = load_workbook(raw_path, data_only=True)
            ws: Worksheet | None = wb.active

            if ws is None:
                raise ValueError("Worksheet non trovato nel workbook")

            # Riga 2 = header
            headers = [str(ws.cell(2, col).value) for col in range(1, ws.max_column + 1)]
            logger.info(f"Headers trovati: {headers}")

            col_map = self._build_column_map(headers)

            records = []
            # Riga 3+ = dati
            for row_idx in range(3, ws.max_row + 1):
                row = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]

                # Skip righe vuote
                if not any(row):
                    continue

                record = self._parse_row(row, col_map, stagione)
                if record:
                    records.append(record)

            logger.info(f"Parsed {len(records)} giocatori da {raw_path}")
            return records

        except Exception as e:
            raise ValueError(f"Failed to parse {raw_path}: {e}") from e

    @staticmethod
    def _build_column_map(headers: list[str]) -> dict[str, int]:
        """Mappa nomi colonne XLSX ai loro indici.

        Colonne XLSX attese: Id, R, Rm, Nome, Squadra, Pv, Mv, Fm, Gf, Gs, Rp, Rc, R+, R-, Ass, Amm, Esp, Au
        """
        col_map = {}
        for idx, header in enumerate(headers):
            h_lower = header.lower().strip()
            if h_lower == "id":
                col_map["id"] = idx
            elif h_lower == "nome":
                col_map["nome"] = idx
            elif h_lower == "squadra":
                col_map["squadra"] = idx
            elif h_lower == "pv":
                col_map["presenze_voto"] = idx
            elif h_lower == "mv":
                col_map["media_voto"] = idx
            elif h_lower == "fm":
                col_map["media_fv_fonte"] = idx
            elif h_lower == "gf":
                col_map["gol"] = idx
            elif h_lower == "gs":
                col_map["gol_subiti"] = idx
            elif h_lower == "rp":
                col_map["rig_parati"] = idx
            elif h_lower == "r+":
                col_map["rig_segnati"] = idx
            elif h_lower == "r-":
                col_map["rig_sbagliati"] = idx
            elif h_lower == "ass":
                col_map["assist_tot"] = idx
            elif h_lower == "amm":
                col_map["ammonizioni"] = idx
            elif h_lower == "esp":
                col_map["espulsioni"] = idx
            elif h_lower == "au":
                col_map["autogol"] = idx

        return col_map

    @staticmethod
    def _parse_row(row: list, col_map: dict, stagione: str) -> dict[str, Any] | None:
        """Parsa una riga di dati da XLSX."""
        try:
            nome_idx = col_map.get("nome")
            squadra_idx = col_map.get("squadra")

            if (
                nome_idx is None
                or squadra_idx is None
                or nome_idx >= len(row)
                or squadra_idx >= len(row)
            ):
                return None

            nome = row[nome_idx]
            squadra = row[squadra_idx]

            if not nome or not squadra:
                return None

            nome = str(nome).strip()
            squadra = str(squadra).strip()

            def get_int(key: str, default: int = 0) -> int:
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return default
                val = row[idx]
                if val is None or val == "":
                    return default
                try:
                    return int(val)
                except (ValueError, TypeError):
                    return default

            def get_float(key: str, default: float = 6.0) -> float:
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return default
                val = row[idx]
                if val is None or val == "":
                    return default
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return default

            record = {
                "id": get_int("id"),
                "nome": nome,
                "squadra": squadra,
                "stagione": stagione,
                "presenze_voto": get_int("presenze_voto"),
                "media_voto": get_float("media_voto"),
                "media_fv_fonte": get_float("media_fv_fonte"),
                "gol": get_int("gol"),
                "assist_tot": get_int("assist_tot"),
                "ammonizioni": get_int("ammonizioni"),
                "espulsioni": get_int("espulsioni"),
                "rig_segnati": get_int("rig_segnati"),
                "rig_parati": get_int("rig_parati"),
                "rig_sbagliati": get_int("rig_sbagliati"),
                "gol_subiti": get_int("gol_subiti"),
                "autogol": get_int("autogol"),
            }

            return record

        except (ValueError, IndexError, TypeError) as e:
            logger.warning(f"Failed to parse row: {e}")
            return None
